import time
import shutil

import deleteBattList
import pandas as pd
from datetime import date
import os
import xlrd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import gc

# Global settings
pd.options.display.max_columns = 8


def xls_to_xlsx(excel_file, xlsx_file):
    """Convert an .xls file to .xlsx format."""
    xls_workbook = xlrd.open_workbook(excel_file)
    xls_sheet = xls_workbook.sheet_by_index(0)
    xlsx_workbook = openpyxl.Workbook()
    xlsx_sheet = xlsx_workbook.active

    for row_index in range(xls_sheet.nrows):
        row = xls_sheet.row_values(row_index)
        xlsx_sheet.append(row)

    xlsx_workbook.save(xlsx_file)


def remove_file(file_path):
    """Remove a file from the filesystem."""
    try:
        os.remove(path=file_path)
    except PermissionError as e:
        print(f"Failed to remove {file_path}. Reason: {e}")


def process_dataframe(wb, sheet_name):
    """Load and process the Excel data."""
    df = pd.read_excel(io=wb, sheet_name=sheet_name, parse_dates=['Battery Replaced Date'])
    df = df[['Building Name', 'Floor Name', 'Room', 'Number Of Batteries', 'Battery Voltage', 'Battery Replaced Date']]
    df.sort_values(by=['Number Of Batteries', 'Battery Voltage', 'Battery Replaced Date'], ascending=True, inplace=True)

    df['Battery Replaced Date'] = pd.to_datetime(df['Battery Replaced Date'], yearfirst=True, errors='coerce')
    df['DaysPassed'] = (pd.Timestamp(date.today()) - df['Battery Replaced Date']).dt.days
    df = df[df['DaysPassed'] > 59]

    return df


def split_and_filter(df):
    """Split and filter the DataFrame based on battery conditions."""
    zerosDF = df[df['Number Of Batteries'] == 0]

    zerosDF = zerosDF[zerosDF['Building Name'] == 'Residential College South']

    foursDF = df[(df['Number Of Batteries'] == 4) & (df['Battery Voltage'] < 5.00) & (df['DaysPassed'] > 89)]

    eightsDF = df[(df['Number Of Batteries'] == 8) & (df['Battery Voltage'] < 8.10) & (df['Battery Voltage'] > 6.25) & (df['DaysPassed'] > 89)]

    return zerosDF, foursDF, eightsDF


def save_dataframe(df, wb, sheet_name):
    """Save DataFrame to an Excel file."""
    df = df[['Building Name', 'Floor Name', 'Room', 'Number Of Batteries', 'Battery Voltage']]
    with pd.ExcelWriter(path=wb, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)


def adjust_column_width(wb):
    """Adjust the column width in an Excel sheet to fit the content."""
    workbook = load_workbook(wb)
    for sheet_name in workbook.sheetnames:
        for column_cells in workbook[sheet_name].columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            if max_length > 0:
                workbook[sheet_name].column_dimensions[column_letter].width = max_length
    workbook.save(wb)
    workbook.close()

def release_resources():
    #force garbage collection to clean up any lingering references
    gc.collect()

def main():
    # Define paths and variables
    dlwb = "C:/Users/HOUSINGADMIN/Downloads/andys.XLS"
    wb = 'C:/Users/HOUSINGADMIN/Downloads/andys.xlsx'
    today = str(date.today())
    new_wb = f'C:/Users/HOUSINGADMIN/Desktop/Low Battery/{today}.xlsx'
    sheet_p1 = 'Sheet'
    sheet_p2 = 'Edits'

    # Convert xls to xlsx
    xls_to_xlsx(dlwb, wb)

    # Remove the original xls file
    remove_file(dlwb)

    # Process the data
    p1DF = process_dataframe(wb, sheet_p1)
    zerosDF, foursDF, eightsDF = split_and_filter(p1DF)

    # Combine the filtered data
    p2DF = pd.concat([zerosDF, foursDF, eightsDF], ignore_index=True)
    p2DF.sort_values(by=['Building Name', 'Floor Name','Room'], ascending=True, inplace=True)

    # Save the processed data
    save_dataframe(p2DF, wb, sheet_p2)

    # Adjust column widths
    adjust_column_width(wb)

    #garbage
    release_resources()

    # Attempt to rename/move the file
    try:
        shutil.copyfile(wb, new_wb)
        time.sleep(5)
        remove_file(wb)
    except PermissionError as e:
        print(f"PermissionError during file operation: {e}")

    print(p2DF)
    quit()


if __name__ == "__main__":
    main(
